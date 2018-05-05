#coding=utf8
_Author_ = 'Alvis'
_Date_ = '2018-04-29 20:56'

import re
import requests
import openpyxl
from time import sleep
import time
from requests.packages import urllib3
class Get_qd_novel:

    def __init__(self):
        self.novel_name = []
        self.novel_author = []
        self.novel_big_type = []
        self.novel_small_type = []
        self.novel_status=[]
        self.novel_intro=[]
        with open("log.text", 'a+') as f:
            f.write("程序启动时间为:"+time.strftime("%Y-%m-%d %H:%M:%S")  + "\n")
            f.close()

    def GetNovelData(self):
        for page in range(1,4000):
            headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36"}
            qd_url = 'https://www.qidian.com/all?orderId=&page=%s&style=1&pageSize=20&siteid=1&pubflag=0&hiddenField=0'%page
            try:
                response = requests.get(qd_url,headers=headers,verify=False,timeout=5)
            except Exception as net_error:
                with open("log.text",'a+') as f:
                    f.write(time.strftime("%Y-%m-%d %H:%M:%S  ")+"第%s页无法获取到数据  "%page +"错误为:%s"%net_error + "\n")
                    f.close()
                continue
            urllib3.disable_warnings()
            html = response.text

            # 小说名称
            name_contents = re.findall('<h4><a href="//book.qidian.com/info/(.*?)/a></h4>', html, re.S)
            # 小说作者
            author_contents = re.findall('<a class="name" href="//my.qidian.com/author/(.*?)/a>', html, re.S)
            # 小说大类型
            big_type_contents = re.findall('</em><a href="//www.(.*?)/a>', html, re.S)
            # 小说小类型
            small_type_contents = re.findall('data-eid="qd_B61">(.*?)</a>', html, re.S)
            # 小说状态
            status_contents = re.findall('</em><span >(.*?)</span>', html, re.S)
            # 小说简介
            intro_contents = re.findall('<p class="intro">(.*?)</p>', html, re.S)

            for i in range(len(intro_contents)):
                self.novel_name.append(re.findall('>(.*?)<', name_contents[i], re.S)[0])
                self.novel_author.append(re.findall('>(.*?)<', author_contents[i], re.S)[0])
                self.novel_big_type.append(re.findall('>(.*?)<', big_type_contents[i], re.S)[0])
                self.novel_small_type.append(small_type_contents[i])
                self.novel_status.append(status_contents[i])
                self.novel_intro.append(intro_contents[i].strip())

            print("已爬取第%s页内容"%page)

    def SaveDataToExcel(self):
        with open("log.text", 'a+') as f:
            f.write("正在写入数据  "+time.strftime("%Y-%m-%d %H:%M:%S")  + "\n")
            f.close()
        try:
            # 存在文件则进行加载
            wb = openpyxl.load_workbook(filename="起点中文网小说名单.xlsx")
        except Exception as e:
            with open("log.text", 'a+') as f:
                f.write("文件名不存在,进行创建  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
                f.close()
            # 不存在则进行创建
            wb = openpyxl.Workbook()
        # 获取所有的表
        all = wb.sheetnames
        # 删除表Sheet
        name = 'Sheet'
        if name in all:
            del wb['Sheet']
        # 创建新表
        ws = wb.create_sheet()
        # 为新表命名
        ws.title ='小说信息'
        ws.cell(row=1, column=1, value='小说名称')
        ws.cell(row=1, column=2, value='小说作者')
        ws.cell(row=1, column=3, value='小说大类型')
        ws.cell(row=1, column=4, value='小说小类型')
        ws.cell(row=1, column=5, value='小说状态')
        ws.cell(row=1, column=6, value='小说简介')

        for i in range(len(self.novel_name)):
            with open("log.text", 'a+') as f:
                f.write("开始写入小说名  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            ws.cell(row=i + 2, column=1, value=self)
            print("已写入%s个小说名" % i)

        for i in range(len(self.novel_author)):
            with open("log.text", 'a+') as f:
                f.write("开始写入小说作者  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            ws.cell(row=i + 2, column=2, value=self.novel_author[i])
            print("已写入%s个作者" % i)

        for i in range(len(self.novel_big_type)):
            with open("log.text", 'a+') as f:
                f.write("开始写入小说大类型  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            ws.cell(row=i + 2, column=3, value=self.novel_big_type[i])
            print("已写入%s个大类型" % i)

        for i in range(len(self.novel_small_type)):
            with open("log.text", 'a+') as f:
                f.write("开始写入小说小类型  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            ws.cell(row=i + 2, column=4, value=self.novel_small_type[i])
            print("已写入%s个小类型" % i)

        for i in range(len(self.novel_status)):
            with open("log.text", 'a+') as f:
                f.write("开始写入小说状态  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            ws.cell(row=i + 2, column=5, value=self.novel_status[i])
            print("已写入%s个状态" % i)

        for i in range(len(self.novel_intro)):
            with open("log.text", 'a+') as f:
                f.write("开始写入小说简介  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            ws.cell(row=i + 2, column=6, value=self.novel_intro[i])
            print("已写入%s个简介"%i)

        try:
            with open("log.text", 'a+') as f:
                f.write("正在保存文件  " + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
            wb.save("起点中文网小说名单.xlsx")
        except Exception as save_fault:
            with open("log.text", 'a+') as f:
                f.write("文件保存失败,原因: %s  "%save_fault + time.strftime("%Y-%m-%d %H:%M:%S") + "\n")
        with open("log.text", 'a+') as f:
            f.write("程序结束时间为:"+time.strftime("%Y-%m-%d %H:%M:%S")  + "\n")
            f.close()



if __name__=="__main__":
    Spider=Get_qd_novel()
    Spider.GetNovelData()
    Spider.SaveDataToExcel()